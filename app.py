import openpyxl
import pyperclip
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
from tqdm import tqdm

driver = webdriver.Chrome()
driver.get('https://cadastro-produtos-devaprender.netlify.app/index.html')
driver.minimize_window()

workbook = openpyxl.load_workbook('auto_xlsx-site_selenium/Produtos_ficticios.xlsx')
sheet_produtos = workbook['Produtos']

def proxima_etapa():
    proxima = driver.find_element(By.XPATH, "//button[@class='btn btn-primary me-2']")
    proxima.click()

def finaliza():
    proxima = driver.find_element(By.XPATH, "//button[@class='btn btn-primary me-2']")
    proxima.click()
    alertas()
    driver.switch_to.default_content()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//button[@class='btn btn-primary']")))
    novo = driver.find_element(By.XPATH, "//button[@class='btn btn-primary']")
    novo.click()

def alertas():
    try:
        WebDriverWait(driver, 1).until(EC.alert_is_present())
        alerta = driver.switch_to.alert
        alerta.accept()
        alerta.accept()
    except WebDriverException:
        pass
    

for linha in tqdm(sheet_produtos.iter_rows(min_row=2), dynamic_ncols=False, total=sheet_produtos.max_row-1, leave=True, unit_scale=True):

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@id='product_name']")))
    pyperclip.copy(linha[0].value)
    input_nome = driver.find_element(By.XPATH, "//input[@id='product_name']")
    input_nome.send_keys(pyperclip.paste())

    pyperclip.copy(linha[1].value)
    input_descricao = driver.find_element(By.XPATH, "//textarea[@id='description']")
    input_descricao.send_keys(pyperclip.paste())

    pyperclip.copy(linha[2].value)
    input_categoria = driver.find_element(By.XPATH, "//input[@id='category']")
    input_categoria.send_keys(pyperclip.paste())

    pyperclip.copy(linha[3].value)
    input_codigo = driver.find_element(By.XPATH, "//input[@id='product_code']")
    input_codigo.send_keys(pyperclip.paste())

    pyperclip.copy(linha[4].value)
    input_peso = driver.find_element(By.XPATH, "//input[@id='weight']")
    input_peso.send_keys(pyperclip.paste())

    pyperclip.copy(linha[5].value)
    input_dimencao = driver.find_element(By.XPATH, "//input[@id='dimensions']")
    input_dimencao.send_keys(pyperclip.paste())
    
    proxima_etapa()

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@id='price']")))
    pyperclip.copy(linha[6].value)
    input_preco = driver.find_element(By.XPATH, "//input[@id='price']")
    input_preco.send_keys(pyperclip.paste())

    pyperclip.copy(linha[7].value)
    input_estoque = driver.find_element(By.XPATH, "//input[@id='stock']")
    input_estoque.send_keys(pyperclip.paste())

    pyperclip.copy(linha[8].value)
    input_validade = driver.find_element(By.XPATH, "//input[@id='expiry_date']")
    input_validade.send_keys(pyperclip.paste())

    pyperclip.copy(linha[9].value)
    input_cor = driver.find_element(By.XPATH, "//input[@id='color']")
    input_cor.send_keys(pyperclip.paste())

    input_tamanho = driver.find_element(By.XPATH, "//select[@id='size']")
    opcao_pequeno = driver.find_element(By.XPATH, "//option[@value='Pequeno']")
    opcao_medio = driver.find_element(By.XPATH, "//option[@value='Médio']")
    opcao_grande = driver.find_element(By.XPATH, "//option[@value='Grande']")

    tamanho = linha[10].value
    input_tamanho.click()
    if tamanho == 'Pequeno':
        opcao_pequeno.click()

    elif tamanho == 'Médio':
        opcao_medio.click()

    else:
        opcao_grande.click()

    pyperclip.copy(linha[11].value)
    input_material = driver.find_element(By.XPATH, "//input[@id='material']")
    input_material.send_keys(pyperclip.paste())

    proxima_etapa()

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@id='manufacturer']")))
    pyperclip.copy(linha[12].value)
    input_fabricante = driver.find_element(By.XPATH, "//input[@id='manufacturer']")
    input_fabricante.send_keys(pyperclip.paste())

    pyperclip.copy(linha[13].value)
    input_origem = driver.find_element(By.XPATH, "//input[@id='country']")
    input_origem.send_keys(pyperclip.paste())

    pyperclip.copy(linha[14].value)
    input_observacoes = driver.find_element(By.XPATH, "//textarea[@id='remarks']")
    input_observacoes.send_keys(pyperclip.paste())

    pyperclip.copy(linha[15].value)
    input_barra = driver.find_element(By.XPATH, "//input[@id='barcode']")
    input_barra.send_keys(pyperclip.paste())

    pyperclip.copy(linha[16].value)
    input_armazem = driver.find_element(By.XPATH, "//input[@id='warehouse_location']")
    input_armazem.send_keys(pyperclip.paste())

    finaliza()

    